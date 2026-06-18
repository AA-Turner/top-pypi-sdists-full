from __future__ import annotations

import base64
import csv
import datetime
import io
import json
import os
import pkgutil
from pathlib import Path
from shutil import move
from tempfile import mkdtemp
from typing import Any, Literal, Optional

import flask
import openpyxl
import pypdfium2 as pdfium

from abstra_internals.cloud_api import (
    get_api_key_info,
    get_feature_flags,
    get_project_info,
)
from abstra_internals.consts.filepaths import TEST_DATA_FILEPATH
from abstra_internals.controllers.execution.drain import (
    drain_until_response,
    normalize_response,
)
from abstra_internals.controllers.linter_events import LinterEventController
from abstra_internals.credentials import (
    delete_credentials,
    get_credentials,
    resolve_headers,
    set_credentials,
)
from abstra_internals.entities.execution_context import (
    HookContext,
    JobContext,
    PageContext,
    Request,
    Response,
    ScriptContext,
)
from abstra_internals.environment import (
    DRAIN_START_TIMEOUT_SECONDS,
    WORKER_LOG_TO_QUEUE,
    web_editor_uses_db,
)
from abstra_internals.interface.cli.deploy import deploy_without_git
from abstra_internals.interface.cli.deploy_messages import DeployMessages
from abstra_internals.interface.contract import ExecutionStartedMessage
from abstra_internals.logger import AbstraLogger
from abstra_internals.repositories.email import EmailRepository
from abstra_internals.repositories.execution import ExecutionFilter, ExecutionRepository
from abstra_internals.repositories.execution_logs import (
    ExecutionLogsRepository,
    LogEntry,
)
from abstra_internals.repositories.factory import Repositories
from abstra_internals.repositories.keyvalue import KVRepository
from abstra_internals.repositories.passwordless import PasswordlessRepository
from abstra_internals.repositories.producer import ProducerRepository
from abstra_internals.repositories.project.project import (
    FormStage,
    HookStage,
    JobStage,
    PageStage,
    ScriptStage,
    Stage,
    StageWithFile,
    StyleSettingsWithSidebar,
)
from abstra_internals.repositories.roles import RolesRepository
from abstra_internals.repositories.tasks import ExecutionTasksResponse, TasksRepository
from abstra_internals.repositories.users import UsersRepository
from abstra_internals.services.fs import FileSystemService
from abstra_internals.services.requirements import RequirementsRepository
from abstra_internals.settings import Settings
from abstra_internals.templates import (
    ensure_dotenv,
    ensure_gitignore,
    new_form_code,
    new_hook_code,
    new_job_code,
    new_page_code,
    new_script_code,
)
from abstra_internals.utils.datetime import to_utc_iso_string
from abstra_internals.utils.file import path2module
from abstra_internals.utils.file_search import (
    find_files_by_glob,
    grep_files,
    list_directory_entries,
)
from abstra_internals.utils.validate import validate_json

MAX_LINES = 500
HARD_MAX_LINES = 50
HARD_MAX_PDF_PAGES = 3

READ_DOCUMENT_MAX_IMAGE_DIMENSION = 1568
READ_DOCUMENT_JPEG_QUALITY = 80


class UnknownNodeTypeError(Exception):
    def __init__(self, node_type: str):
        self.node_type = node_type

    def __str__(self):
        return f"Unknown node type {self.node_type}"


class SelfTransitionError(Exception):
    def __init__(self, node_type: str, node_id: str):
        self.node_type = node_type
        self.node_id = node_id

    def __str__(self):
        return "You can't add a transition to itself."


class TransitionToJobError(Exception):
    def __init__(
        self, source_type: str, source_id: str, target_type: str, target_id: str
    ):
        self.source_type = source_type
        self.source_id = source_id
        self.target_type = target_type
        self.target_id = target_id

    def __str__(self):
        return "You can't add a transition to a job. Use a script instead."


class DoubleTransitionError(Exception):
    def __init__(
        self, source_type: str, source_id: str, target_type: str, target_id: str
    ):
        self.source_type = source_type
        self.source_id = source_id
        self.target_type = target_type
        self.target_id = target_id

    def __str__(self):
        return "You can't add the same transition twice."


class MainController:
    kv_repository: KVRepository
    passwordless_repository: PasswordlessRepository
    email_repository: EmailRepository
    tasks_repository: TasksRepository
    users_repository: UsersRepository
    roles_repository: RolesRepository
    producer_repository: ProducerRepository
    execution_repository: ExecutionRepository
    execution_logs_repository: ExecutionLogsRepository

    def __init__(self, repositories: Repositories):
        repositories.project.initialize_or_migrate()

        RequirementsRepository.ensure("abstra")
        ensure_gitignore(Settings.root_path)
        ensure_dotenv(Settings.root_path)

        self.repositories = repositories

        self.kv_repository = repositories.kv
        self.passwordless_repository = repositories.passwordless
        self.email_repository = repositories.email
        self.users_repository = repositories.users
        self.roles_repository = repositories.roles
        self.tasks_repository = repositories.tasks
        self.producer_repository = repositories.producer
        self.execution_repository = repositories.execution
        self.execution_logs_repository = repositories.execution_logs
        self.linter_repository = repositories.linter
        self.code_markers_repository = repositories.code_markers

    def deploy_without_git(self):
        DeployMessages.start(method="upload")
        DeployMessages.checking_linters()

        issues = self.linter_repository.get_blocking_checks_for_deploy()
        LinterEventController.broadcast(self.linter_repository.checks)

        if len(issues) > 0:
            raise Exception(
                "Please fix all linter issues before deploying your project."
            )

        deploy_without_git(show_start_message=False)

    def reset_execution_repository(self):
        self.execution_repository.clear()

    def reset_execution_logs_repository(self):
        self.execution_logs_repository.clear()

    def reset_tasks_repository(self):
        self.tasks_repository.clear()

    def reset_repositories(self):
        self.reset_execution_logs_repository()
        self.reset_execution_repository()
        self.reset_tasks_repository()

    def get_workspace(self) -> StyleSettingsWithSidebar:
        """
        Get the current workspace settings including styling and sidebar configuration.

        Retrieves the project's branding settings and navigation sidebar items.

        Copywritings:
            Get workspace settings
            Retrieving workspace settings...
        """
        project = self.repositories.project.load()
        return project.get_workspace()

    def get_stage(self, id: str) -> Stage | None:
        """
        Retrieve a specific workflow stage by its unique identifier.

        This method looks up and returns a single stage from the project
        based on the provided ID. The stage can be of any type (form, hook,
        job, or tasklet).

        Args:
            id (str): Unique identifier of the stage to retrieve.


        Note:
            - Returns None if no stage with the given ID exists
            - The ID is case-sensitive and must match exactly
            - Use isinstance() to determine the specific stage type

        Copywritings:
            Get a specific workflow stage
            Retrieving a specific workflow stage by ID...
        """
        project = self.repositories.project.load()
        return project.get_stage(id)

    @staticmethod
    def _safe_mtime(file_path: Path) -> float:
        try:
            return file_path.stat().st_mtime
        except OSError:
            return 0.0

    def _read_file_lines_with_pagination(
        self,
        file_path: Path,
        start_line: int | None,
        end_line: int | None,
        max_lines: int,
        encoding: str = "utf-8",
    ) -> dict[str, Any] | None:
        """
        Private helper method to read file lines with pagination support.

        This method contains the common pagination logic used by both
        read_stage_file_with_pagination and read_file_with_pagination.

        Args:
            file_path (Path): Absolute path to the file to read.
            start_line (Optional[int]): 1-indexed line number to start reading from.
            end_line (Optional[int]): 1-indexed line number to stop reading at (inclusive).
            max_lines (int): Maximum number of lines to return in a single call.
            encoding (str): File encoding. Defaults to "utf-8".

        Returns:
            dict | None: Dictionary containing file content and metadata, or None if file cannot be read.
        """
        try:
            relative_file = (
                file_path.resolve().relative_to(Settings.root_path.resolve()).as_posix()
            )
        except ValueError:
            return {"error": "file must be a path inside the project"}

        if not file_path.is_file():
            return None

        # First pass: count total lines without loading into memory
        try:
            with file_path.open("r", encoding=encoding) as f:
                total_lines = sum(1 for _ in f)
        except UnicodeDecodeError:
            return {
                "error": f"Failed to decode file with encoding '{encoding}'. Try passing encoding: 'latin-1' as a parameter."
            }
        except LookupError:
            return {
                "error": f"Unknown encoding '{encoding}'. Common encodings: 'utf-8', 'latin-1', 'ascii', 'utf-16'."
            }
        except OSError as e:
            AbstraLogger.error(f"Failed to read file {file_path}: {e}")
            return None

        # Determine actual range to read
        actual_start = max(1, start_line or 1)
        actual_end = min(total_lines, end_line or total_lines)

        # Handle edge case where start_line exceeds total_lines
        if actual_start > total_lines:
            return {
                "content": "",
                "start_line": actual_start,
                "end_line": actual_start - 1,
                "total_lines": total_lines,
                "has_more": False,
                "truncated": False,
                "mtime": self._safe_mtime(file_path),
                "file": relative_file,
            }

        # Apply max_lines limit
        truncated = False
        if actual_end - actual_start + 1 > max_lines:
            actual_end = actual_start + max_lines - 1
            truncated = True

        # Second pass: read only the required line range using itertools.islice
        from itertools import islice

        try:
            with file_path.open("r", encoding=encoding) as f:
                # Skip lines before start (0-indexed), then take the range we need
                start_idx = actual_start - 1
                num_lines = actual_end - actual_start + 1
                content_lines = list(islice(f, start_idx, start_idx + num_lines))
                content = "".join(content_lines)
        except UnicodeDecodeError:
            return {
                "error": f"Failed to decode file with encoding '{encoding}'. Try passing encoding: 'latin-1' as a parameter."
            }
        except LookupError:
            return {
                "error": f"Unknown encoding '{encoding}'. Common encodings: 'utf-8', 'latin-1', 'ascii', 'utf-16'."
            }
        except OSError as e:
            AbstraLogger.error(f"Failed to read file {file_path}: {e}")
            return None

        return {
            "content": content,
            "start_line": actual_start,
            "end_line": actual_end,
            "total_lines": total_lines,
            "has_more": actual_end < total_lines,
            "truncated": truncated,
            "mtime": self._safe_mtime(file_path),
            "file": relative_file,
        }

    def read_stage_file_with_pagination(
        self,
        id: str,
        start_line: int | None = None,
        end_line: int | None = None,
        max_lines: int = 500,
    ):
        """
        Read the source code of a stage's Python file by stage ID, with pagination.

        Use this when you know the stage ID. Use list_all_stages to find stage IDs.
        For reading non-stage files, use read_file_with_pagination instead.

        Args:
            id (str): Unique identifier of the stage whose file to read.
            start_line (Optional[int]): 1-indexed line number to start reading from.
                If None, starts from the beginning. Defaults to None.
            end_line (Optional[int]): 1-indexed line number to stop reading at (inclusive).
                If None, reads until max_lines is reached. Defaults to None.
            max_lines (int): Maximum number of lines to return in a single call.
                Hard maximum of 1000. Defaults to 500.

        Example:
            ```python
            controller = MainController(repositories)

            # Read entire small stage file
            result = controller.read_stage_file_with_pagination("script-456")
            if result:
                print(result["content"])
                print(f"File has {result['total_lines']} lines")

            # Read first 200 lines of a large stage file
            result = controller.read_stage_file_with_pagination("form-123", start_line=1, end_line=200)
            if result:
                print(result["content"])
                if result["has_more"]:
                    print("Stage file has more content...")

            # Read specific lines 50-250
            result = controller.read_stage_file_with_pagination("hook-789", start_line=50, end_line=250)
            ```

        Note:
            - Returns None if the stage does not exist or has no file
            - The ID is case-sensitive and must match exactly
            - Line numbers are 1-indexed (first line is 1, not 0)
            - Automatic truncation at max_lines prevents context overflow
            - For files with more than max_lines, use pagination to read chunks

        Copywritings:
            Read stage file with pagination
            Reading stage file with pagination support...
        """
        HARD_MAX_LINES_LIMIT = 1000
        if max_lines > HARD_MAX_LINES_LIMIT:
            return {
                "error": f"max_lines cannot exceed {HARD_MAX_LINES_LIMIT}, got {max_lines}"
            }

        stage = self.get_stage(id)
        if not isinstance(stage, StageWithFile):
            return None

        return self._read_file_lines_with_pagination(
            stage.file_path, start_line, end_line, max_lines
        )

    def get_async_stage_ids(self):
        project = self.repositories.project.load()
        job_ids = [stage.id for stage in project.jobs]
        script_ids = [stage.id for stage in project.scripts]
        return job_ids + script_ids

    def __ensure_case(self, path: str):
        file_dirs = [p for p in Settings.root_path.iterdir()]
        if path in file_dirs:
            return

        conflicting_paths = [p for p in file_dirs if p.name.lower() == path.lower()]
        if len(conflicting_paths) == 1:
            conflicting_paths[0].rename(Settings.root_path.joinpath(path))
            return

        raise Exception(
            f"File {path} already exists with different casing. Conflict with files ({', '.join(p.name for p in conflicting_paths)})"
        )

    def init_code_file(self, path: str, code: str):
        file_path = Settings.root_path.joinpath(path)
        if file_path.is_file():
            self.__ensure_case(path)
            return
        file_path.parent.mkdir(parents=True, exist_ok=True)
        file_path.write_text(code, encoding="utf-8")

    def read_file(self, file: str):
        """
        Read the contents of a file from the project workspace.

        This method reads and returns the text content of a file within the project
        directory. It handles encoding properly and returns None if the file doesn't exist.

        Args:
            file (str): Relative path to the file from the project root directory.
                Should include the file extension.

        Returns:
            str | None: The file content as a string if the file exists and is readable,
                None if the file doesn't exist or is not a regular file.

        Note:
            - Files are read with UTF-8 encoding
            - Path should be relative to the project root directory
            - Returns None for directories, non-existent files, or unreadable files
            - Suitable for text files; binary files may not be read correctly
            - Does not raise exceptions for missing files, returns None instead

        Copywritings:
            Read the contents of a file
            Reading the contents of a file...
        """
        file_path = Settings.root_path.joinpath(file)
        if not file_path.is_file():
            return None
        return Settings.root_path.joinpath(file).read_text(encoding="utf-8")

    def _read_spreadsheet_file(
        self,
        file_path: Path,
        sheet_name: str | None = None,
        start_line: int | None = None,
        end_line: int | None = None,
    ) -> dict[str, Any] | None:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            AbstraLogger.error(f"Failed to read spreadsheet {file_path}: {e}")
            return None

        if not sheet_name:
            summary_lines = []
            for name in wb.sheetnames:
                ws = wb[name]
                rows = ws.max_row or 0
                cols = ws.max_column or 0
                summary_lines.append(f"- {name}: {rows} rows, {cols} columns")
            sheets = list(wb.sheetnames)
            wb.close()
            return {
                "content": "Sheets in this file:\n"
                + "\n".join(summary_lines)
                + "\n\nUse sheet_name parameter to read a specific sheet.",
                "type": "spreadsheet_summary",
                "sheets": sheets,
            }

        if sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            wb.close()
            return {
                "content": f'Error: Sheet "{sheet_name}" not found. Available sheets: {available}',
                "error": True,
            }

        ws = wb[sheet_name]
        total_rows = ws.max_row or 0

        actual_start = max(1, start_line or 1)
        actual_end = min(total_rows, end_line or total_rows)

        if actual_start > total_rows:
            wb.close()
            return {
                "content": "",
                "start_line": actual_start,
                "end_line": actual_start - 1,
                "total_lines": total_rows,
                "has_more": False,
                "truncated": False,
                "sheet_name": sheet_name,
            }

        truncated = False
        if actual_end - actual_start + 1 > HARD_MAX_LINES:
            actual_end = actual_start + HARD_MAX_LINES - 1
            truncated = True

        # Only iterate the rows we need
        output = io.StringIO()
        writer = csv.writer(output)
        for row in ws.iter_rows(
            min_row=actual_start, max_row=actual_end, values_only=True
        ):
            cells = [cell if cell is not None else "" for cell in row]
            while cells and cells[-1] == "":
                cells.pop()
            if cells:
                writer.writerow(cells)
        wb.close()

        content = output.getvalue().rstrip("\n")

        return {
            "content": content,
            "start_line": actual_start,
            "end_line": actual_end,
            "total_lines": total_rows,
            "has_more": actual_end < total_rows,
            "truncated": truncated,
            "sheet_name": sheet_name,
        }

    def _read_pdf_file(
        self,
        file_path: Path,
        start_page: int | None = None,
        end_page: int | None = None,
    ) -> dict[str, Any] | None:
        try:
            doc = pdfium.PdfDocument(str(file_path))
            total_pages = len(doc)

            actual_start = max(1, start_page or 1)
            requested_end = end_page or actual_start + HARD_MAX_PDF_PAGES - 1
            actual_end = min(
                requested_end,
                total_pages,
                actual_start + HARD_MAX_PDF_PAGES - 1,
            )

            pages_text = []
            for i in range(actual_start - 1, actual_end):
                page = doc[i]
                text_page = page.get_textpage()
                text = text_page.get_text_range()
                pages_text.append(f"--- Page {i + 1} ---\n{text}")
                text_page.close()

            doc.close()

            has_more = actual_end < total_pages
            content = "\n\n".join(pages_text)
            content += (
                f"\n\n[Pages {actual_start}-{actual_end} of {total_pages}"
                + (", use start_page/end_page to read more" if has_more else "")
                + "]"
            )

            return {
                "content": content,
                "total_pages": total_pages,
                "start_page": actual_start,
                "end_page": actual_end,
                "has_more": has_more,
            }
        except Exception as e:
            AbstraLogger.error(f"Failed to read PDF {file_path}: {e}")
            return None

    def _downscale_image_to_jpeg(self, file_path: Path) -> bytes | None:
        """Downscale + JPEG re-encode to cap token cost. Returns None if the
        image can't be decoded, so the caller can fall back to the raw bytes."""
        try:
            from PIL import Image

            from abstra_internals.utils.image import constrain_image_size

            with Image.open(file_path) as img:
                img.load()
                img = constrain_image_size(
                    img, max_dimension=READ_DOCUMENT_MAX_IMAGE_DIMENSION
                )
                if img.mode not in ("RGB", "L"):
                    img = img.convert("RGB")
                buffer = io.BytesIO()
                img.save(buffer, format="JPEG", quality=READ_DOCUMENT_JPEG_QUALITY)
                return buffer.getvalue()
        except Exception as e:
            AbstraLogger.warning(f"Could not downscale image {file_path}: {e}")
            return None

    def _read_image_file(
        self,
        file_path: Path,
    ) -> dict[str, Any] | None:
        try:
            jpeg_bytes = self._downscale_image_to_jpeg(file_path)
            if jpeg_bytes is not None:
                mime_type = "image/jpeg"
                image_data = base64.b64encode(jpeg_bytes).decode("utf-8")
            else:
                # Fallback: pass the original bytes through unchanged.
                ext = file_path.suffix.lower()
                mime_type = {
                    ".png": "image/png",
                    ".gif": "image/gif",
                    ".webp": "image/webp",
                }.get(ext, "image/jpeg")
                image_data = base64.b64encode(file_path.read_bytes()).decode("utf-8")

            data_uri = f"data:{mime_type};base64,{image_data}"

            return {
                "__imageContent": True,
                "mimeType": mime_type,
                "dataUri": data_uri,
                "description": f"Image file: {file_path.name}",
            }
        except Exception as e:
            AbstraLogger.error(f"Failed to read image {file_path}: {e}")
            return None

    def read_file_with_pagination(
        self,
        file: str,
        start_line: int | None = None,
        end_line: int | None = None,
        max_lines: int = MAX_LINES,
        encoding: str = "utf-8",
    ):
        """
        Read text content of a code or text file with line-range pagination.

        Use this to read project code and text files. For documents (spreadsheets,
        PDFs, images, CSVs), use read_document instead. For finding files by name, use
        find_files_by_pattern. For searching inside files, use search_file_with_context
        (single file) or grep_codebase (all files). For listing directory contents, use
        list_directory.

        Args:
            file (str): Relative path to the file from the project root directory.
                Should include the file extension.
            start_line (Optional[int]): 1-indexed line number to start reading from.
                If None, starts from the beginning. Defaults to None.
            end_line (Optional[int]): 1-indexed line number to stop reading at (inclusive).
                If None, reads until max_lines is reached. Defaults to None.
            max_lines (int): Maximum number of lines to return in a single call.
                Hard maximum of 1000. Defaults to 500.
            encoding (str): File encoding. Defaults to "utf-8".

        Example:
            ```python
            controller = MainController(repositories)

            # Read entire small file
            result = controller.read_file_with_pagination("config.json")
            if result:
                print(result["content"])
                print(f"File has {result['total_lines']} lines")

            # Read first 200 lines of a large file
            result = controller.read_file_with_pagination("large_log.txt", start_line=1, end_line=200)
            if result:
                print(result["content"])
                if result["has_more"]:
                    print("File has more content...")

            # Read lines 500-900
            result = controller.read_file_with_pagination("data.py", start_line=500, end_line=900)
            ```

        Note:
            - Path should be relative to the project root directory
            - Returns None for directories or non-existent files
            - Returns an error dict for encoding failures
            - Line numbers are 1-indexed (first line is 1, not 0)
            - If start_line > total lines, returns empty content with metadata
            - Automatic truncation at max_lines prevents context overflow
            - For files with more than max_lines, use pagination to read chunks

        Copywritings:
            Read file contents
            Reading file contents...
            Reading {file}...
        """
        HARD_MAX_LINES_LIMIT = 1000
        if max_lines > HARD_MAX_LINES_LIMIT:
            return {
                "error": f"max_lines cannot exceed {HARD_MAX_LINES_LIMIT}, got {max_lines}"
            }

        file_path = Settings.root_path.joinpath(file)

        if not file_path.exists():
            return None

        return self._read_file_lines_with_pagination(
            file_path, start_line, end_line, max_lines, encoding
        )

    def read_document(
        self,
        file: str,
        start_line: int | None = None,
        end_line: int | None = None,
        sheet_name: str | None = None,
        start_page: int | None = None,
        end_page: int | None = None,
        encoding: str = "utf-8",
    ):
        """
        Read a document file (spreadsheets, PDFs, images, CSVs, or other non-code files).

        Use this to read documents. It supports:
        - Spreadsheets (xlsx/xls): call without sheet_name to get a summary of sheets,
          with sheet_name to get CSV content of that sheet (paginated by lines, hard limit: 50 lines per call)
        - PDFs: renders pages as images, page-based pagination with start_page/end_page (hard limit: 3 pages per call)
        - Images: returns the image data for visual inspection
        - Other files: reads as text with line-based pagination (hard limit: 50 lines per call)

        For reading project code files, use read_file_with_pagination instead.

        Args:
            file (str): Relative path to the file from the project root directory.
                Should include the file extension.
            start_line (Optional[int]): 1-indexed line number for text/spreadsheet pagination.
                If None, starts from the beginning. Defaults to None.
            end_line (Optional[int]): 1-indexed end line for text/spreadsheet pagination.
                Hard limit: 50 lines per call.
            sheet_name (Optional[str]): Sheet name for xlsx/xls files. Omit to get summary.
            start_page (Optional[int]): 1-indexed start page for PDFs.
                If None, starts from page 1. Defaults to None.
            end_page (Optional[int]): 1-indexed end page for PDFs.
                Hard limit: 3 pages per call.
            encoding (str): File encoding. Defaults to "utf-8".

        Copywritings:
            Read document
            Reading document...
            Reading {file}...
        """
        file_path = Settings.root_path.joinpath(file)

        if not file_path.exists():
            return {"error": f"File not found at path: {file}"}

        suffix = file_path.suffix.lower()

        if suffix in (".jpg", ".jpeg", ".png", ".gif", ".webp"):
            return self._read_image_file(file_path)
        elif suffix in (".xlsx", ".xls"):
            return self._read_spreadsheet_file(
                file_path, sheet_name, start_line, end_line
            )
        elif suffix == ".pdf":
            return self._read_pdf_file(file_path, start_page, end_page)
        else:
            return self._read_file_lines_with_pagination(
                file_path, start_line, end_line, HARD_MAX_LINES, encoding
            )

    def check_file_exists(self, file_path: str):
        """
        Check if a file exists in the project workspace.

        This method verifies whether a file exists at the specified path relative
        to the project root directory. It only returns True for actual files,
        not directories.

        Args:
            file_path (str): Relative path to the file from the project root directory.
                Should include the file extension.

        Returns:
            bool: True if the file exists and is a regular file, False otherwise.


        Note:
            - Only returns True for regular files, not directories or special files
            - Path should be relative to the project root directory
            - Does not check file permissions or readability, only existence
            - Use this before calling read_file() to avoid None returns
            - Faster than read_file() for existence checks

        Copywritings:
            Check if a file exists
            Checking if a file exists...
        """
        return Settings.root_path.joinpath(file_path).is_file()

    def list_files(self, path: str = ".", mode: str = "file"):
        """
        List files and directories within the project workspace.

        This method provides different listing modes to browse the project filesystem,
        supporting various file types and Python module discovery.

        Args:
            path (str, optional): Relative path from project root to list contents.
                Defaults to "." (project root).
            mode (str, optional): Listing mode that determines what to return:
                - "file": All files and directories (default)
                - "image": Only image files (png, jpg, jpeg, gif, svg, webp, etc.)
                - "python-file": Only Python files (.py extension)
                - "module": Python modules and packages discoverable by import

        Returns:
            List[Dict]: List of file/directory information. Structure depends on mode:
                For "file", "image", "python-file" modes:
                - name: Filename relative to the specified path
                - path: Full path relative to project root
                - type: "file" or "dir"

                For "module" mode:
                - name: Module/package name
                - path: Import path (dot notation)
                - type: "module" or "package"

        Example:
            ```python
            controller = MainController(repositories)

            # List all files in project root
            all_files = controller.list_files()
            for item in all_files:
                print(f"{item['type']}: {item['name']} -> {item['path']}")

            # List only Python files in tasklets directory
            python_files = controller.list_files("tasklets", mode="python-file")
            for file in python_files:
                if file['type'] == 'file':
                    print(f"Python file: {file['path']}")

            # List only image files
            images = controller.list_files(mode="image")
            for img in images:
                print(f"Image: {img['name']} ({img['path']})")

            # Discover Python modules
            modules = controller.list_files(mode="module")
            for mod in modules:
                print(f"{mod['type']}: {mod['name']} -> import {mod['path']}")

            # List contents of specific subdirectory
            subdir_files = controller.list_files("forms")
            ```

        Note:
            - Does not filter by .gitignore
            - Image mode supports: .png, .jpg, .jpeg, .gif, .svg, .webp, .jfif, .pjp, .pjpeg
            - Module mode uses Python's pkgutil.iter_modules for discovery
            - Paths are always relative to the project root directory
            - Returns empty list if the specified path doesn't exist

        Copywritings:
            List files and directories in the project workspace
            Retrieving files and directories in the project workspace...
        """
        parent_path = Settings.root_path.joinpath(path)
        if mode in ["file", "image", "python-file"]:
            if mode == "image":
                allowed_suffixes = [
                    ".png",
                    ".jpg",
                    ".jpeg",
                    ".gif",
                    ".svg",
                    ".webp",
                    ".jfif",
                    ".pjp",
                    ".pjpeg",
                ]
            elif mode == "python-file":
                allowed_suffixes = [".py"]
            else:
                allowed_suffixes = None

            return [
                dict(
                    name=str(file.relative_to(parent_path)),
                    path=str(file.relative_to(Settings.root_path)),
                    type="file" if file.is_file() else "dir",
                )
                for file in FileSystemService.list_files(
                    parent_path, allowed_suffixes=allowed_suffixes, use_ignore=False
                )
            ]

        elif mode == "module":
            return [
                dict(
                    name=name,
                    path=path2module(Path(path).joinpath(name)),
                    type="package" if ispkg else "module",
                )
                for module_finder, name, ispkg in pkgutil.iter_modules(
                    [str(parent_path)]
                )
            ]

    def search_file_with_context(
        self,
        file: str,
        pattern: str,
        context_lines: int = 5,
        case_sensitive: bool = True,
        max_matches: int = 50,
    ):
        """
        Search for a regex pattern in a single file with surrounding context lines.

        This is a single-file search with context. Use grep_codebase to search across
        ALL files. Use find_files_by_pattern to find files by name. Use
        read_file_with_pagination to read a known file.

        Supports full Python regex syntax (e.g., r"def \\w+\\(", r"TODO|FIXME").

        Args:
            file (str): Relative path to the file from the project root directory.
            pattern (str): Regular expression pattern to search for.
            context_lines (int): Number of lines to include before and after each match.
                Defaults to 5.
            case_sensitive (bool): Whether the search should be case-sensitive.
                Defaults to True.
            max_matches (int): Maximum number of matches to return to prevent
                context overflow. Defaults to 50.

        Example:
            ```python
            controller = MainController(repositories)

            # Search for function definitions with context
            result = controller.search_file_with_context(
                "utils.py",
                r"def \\w+\\(",
                context_lines=3
            )
            if result:
                print(f"Found {result['total_matches']} matches in {result['file']}")
                for match in result["matches"]:
                    print(f"\nMatch at line {match['match_line']}:")
                    print(match["context"])
                    print("-" * 40)

            # Case-insensitive search for error handling
            result = controller.search_file_with_context(
                "app.py",
                r"error|exception|fail",
                context_lines=5,
                case_sensitive=False
            )

            # Search for specific class with 10 lines of context
            result = controller.search_file_with_context(
                "models.py",
                r"class UserModel",
                context_lines=10,
                max_matches=10
            )

            # Search imports in a large file
            result = controller.search_file_with_context(
                "main.py",
                r"^import |^from .* import",
                context_lines=0  # Just the matching lines
            )
            ```

        Note:
            - Context windows may overlap for nearby matches
            - Line numbers are 1-indexed (first line is 1, not 0)
            - Returns None for non-existent or unreadable files
            - Much more efficient than reading entire file when searching
            - Use max_matches to control context window usage

        Copywritings:
            Search file with context
            Searching file for pattern...
            Searching {file} for '{pattern}'...
        """
        import re

        file_path = Settings.root_path.joinpath(file)
        if not file_path.is_file():
            return None

        # Compile regex pattern
        flags = 0 if case_sensitive else re.IGNORECASE
        try:
            regex = re.compile(pattern, flags)
        except re.error as e:
            AbstraLogger.error(f"Invalid regex pattern '{pattern}': {e}")
            return None

        # First pass: find matching lines and count total without loading all into memory
        matches = []
        total_lines = 0
        try:
            with file_path.open("r", encoding="utf-8") as f:
                for line_num, line in enumerate(f, start=1):
                    total_lines = line_num
                    if regex.search(line):
                        matches.append((line_num, line))
                        # Early exit if we have enough matches for return
                        if len(matches) > max_matches + 100:  # Keep some buffer
                            break
        except OSError as e:
            AbstraLogger.error(f"Failed to read file {file_path}: {e}")
            return None
        except UnicodeDecodeError as e:
            AbstraLogger.error(f"Failed to decode file {file_path}: {e}")
            return None

        # If we stopped early, count remaining lines
        if len(matches) > max_matches + 100:
            try:
                with file_path.open("r", encoding="utf-8") as f:
                    total_lines = sum(1 for _ in f)
            except (OSError, UnicodeDecodeError):
                pass  # Use the line count we have

        total_matches = len(matches)
        truncated = total_matches > max_matches
        matches_to_return = matches[:max_matches]

        # Build context for each match (need to re-read relevant sections)
        from itertools import islice

        result_matches = []
        for match_line, match_text in matches_to_return:
            # Calculate context range
            start_line = max(1, match_line - context_lines)
            end_line = min(total_lines, match_line + context_lines)

            # Read only the context lines needed
            try:
                with file_path.open("r", encoding="utf-8") as f:
                    start_idx = start_line - 1
                    num_lines = end_line - start_line + 1
                    context_lines_list = list(
                        islice(f, start_idx, start_idx + num_lines)
                    )
                    context = "".join(context_lines_list)
            except (OSError, UnicodeDecodeError) as e:
                AbstraLogger.error(f"Failed to read context for {file_path}: {e}")
                context = match_text

            result_matches.append(
                {
                    "match_line": match_line,
                    "match_text": match_text.rstrip("\n"),
                    "start_line": start_line,
                    "end_line": end_line,
                    "context": context,
                }
            )

        return {
            "file": file,
            "total_lines": total_lines,
            "total_matches": total_matches,
            "matches_returned": len(result_matches),
            "truncated": truncated,
            "matches": result_matches,
        }

    def list_directory(self, path: str = "."):
        """
        List immediate contents of a directory (non-recursive, one level).

        Use this to explore project structure step by step. Start with path="."
        for the root. For recursive file search by name, use find_files_by_pattern
        instead. For reading file contents, use read_file_with_pagination.

        Args:
            path (str): Relative path from project root. Defaults to "." (root).


        Note:
            - Only lists *immediate* children, not recursive.
            - Directories are listed before files, both sorted alphabetically.

        Copywritings:
            List directory contents
            Listing directory...
            Listing {path}...
        """
        dir_path = Settings.root_path / path
        if not dir_path.is_dir():
            return None
        entries = list_directory_entries(dir_path, is_ignored_fn=None)
        return {"path": path, "entries": entries}

    def find_files_by_pattern(self, pattern: str, max_results: int = 200):
        """
        Find files by name/path using glob patterns (recursive).

        Use this to locate files when you know part of the name. For searching
        file CONTENTS, use grep_codebase instead. For reading a known file,
        use read_file_with_pagination. Supports ``**`` for recursive matching.
        A bare word (e.g. ``"utils"``) auto-expands to ``"**/utils*"``.

        Args:
            pattern (str): Glob pattern relative to the project root.
                Examples:
                - "**/*.py"       all Python files (recursive)
                - "src/**/*.ts"   TypeScript files under src/
                - "utils"         any file whose name starts with "utils"
                - "**/test_*"     any file starting with "test_"
            max_results (int): Upper bound on results returned. Defaults to 200.

        Example:
            ```python
            # Find all Python files
            py_files = controller.find_files_by_pattern("**/*.py")

            # Find a specific module anywhere in the project
            matches = controller.find_files_by_pattern("file_search")
            # → ["abstra_internals/utils/file_search.py"]

            # Find test files
            test_files = controller.find_files_by_pattern("**/test_*")
            ```

        Note:
            - Only returns files, never directories.
            - Results are capped by max_results to protect context window.

        Copywritings:
            Find files by pattern
            Searching for files...
            Searching for files matching '{pattern}'...
        """
        return find_files_by_glob(
            Settings.root_path,
            pattern,
            is_ignored_fn=None,
            max_results=max_results,
        )

    def grep_codebase(
        self,
        query: str,
        file_pattern: str = "**/*.py",
        case_sensitive: bool = True,
        max_results: int = 100,
    ):
        """
        Search for a string or regex across ALL project files matching a glob.

        This is the cross-file search tool. Use search_file_with_context when you
        already know which file to search (it provides richer context with surrounding
        lines). Use find_files_by_pattern to find files by name/path pattern.
        Results are capped to protect the context window.

        The query is treated as a Python regex.  If it is not a valid regex
        it falls back to a literal string search, so plain text always works.

        Args:
            query (str): Plain text or Python regex to search for.
            file_pattern (str): Glob pattern selecting which files to search.
                Defaults to "**/*.py" (all Python files).
                Use "**/*" to search every file in the project.
            case_sensitive (bool): Whether matching is case-sensitive.
                Defaults to True.
            max_results (int): Maximum number of matching lines returned
                across all files combined. Defaults to 100.

        Example:
            ```python
            # Find all uses of a function across the codebase
            hits = controller.grep_codebase("my_function")

            # Find TODO comments in any file
            todos = controller.grep_codebase(
                "TODO", file_pattern="**/*", case_sensitive=False
            )

            # Find a specific import in Python files
            hits = controller.grep_codebase(
                r"^from abstra import", file_pattern="**/*.py"
            )
            ```

        Note:
            - Binary files are silently skipped.
            - Use search_file_with_context once you know the file, as it
              provides richer context (surrounding lines).

        Copywritings:
            Search across codebase
            Searching codebase...
            Searching for '{query}' in {file_pattern}...
        """
        return grep_files(
            Settings.root_path,
            query,
            file_pattern=file_pattern,
            case_sensitive=case_sensitive,
            max_results=max_results,
            is_ignored_fn=None,
        )

    def update_workspace(self, changes: dict[str, Any]):
        """
        Update workspace branding and styling settings.

        Modifies the project's visual branding including name, logo, colors,
        and language preferences. Changes are saved to abstra.json.

        Args:
            changes (Dict[str, Any]): Dictionary of settings to update.
                Supported keys:
                - name (str): Workspace display name
                - language (str): Language code ('en', 'pt', 'de', 'es', 'fr', 'hi')
                - theme (str): Theme identifier or null
                - logo_url (str): Path to logo image (e.g., './logo.png')
                - favicon_url (str): Path to favicon
                - brand_name (str): Brand display name
                - main_color (str): Primary color hex (e.g., '#FF5733')
                - font_family (str): Font family name
                - font_color (str): Font color hex

        Example:
            ```python
            controller.update_workspace({
                "brand_name": "My Company",
                "main_color": "#3B82F6",
                "language": "en"
            })
            ```

        Copywritings:
            Update workspace branding settings
            Updating workspace branding settings...
        """
        with self.repositories.project.atomic() as project:
            project.workspace.update(changes)
        return project.workspace

    def is_initial(self, id: str):
        project = self.repositories.project.load()
        stage = project.get_stage(id)
        if not stage:
            raise Exception(f"Stage {id} not found")
        return project.is_initial(stage)

    def get_scripts(self) -> list[ScriptStage]:
        project = self.repositories.project.load()
        scripts = project.get_scripts()

        sorted_scripts = sorted(scripts, key=lambda s: s.title.lower())
        return sorted_scripts

    def get_script(self, id: str) -> ScriptStage | None:
        project = self.repositories.project.load()
        return project.get_script(id)

    def get_forms(self) -> list[FormStage]:
        project = self.repositories.project.load()
        forms = project.get_forms()

        sorted_forms = sorted(forms, key=lambda f: f.title.lower())
        return sorted_forms

    def get_form(self, id: str) -> FormStage | None:
        project = self.repositories.project.load()
        return project.get_form(id)

    def get_form_by_path(self, path: str) -> FormStage | None:
        project = self.repositories.project.load()
        return project.get_form_by_path(path)

    def write_test_data(self, data: str) -> None:
        if not validate_json(data):
            raise Exception("Invalid JSON")
        test_file = Settings.root_path / TEST_DATA_FILEPATH
        test_file.write_text(data, encoding="utf-8")

    def read_test_data(self) -> str:
        test_file = Settings.root_path / TEST_DATA_FILEPATH
        if not test_file.is_file():
            return "{}"
        return test_file.read_text(encoding="utf-8")

    def delete_stage(self, stage_id: str, remove_file: bool = False):
        """
        Delete a stage from the project workflow.

        This method removes a stage from the project configuration and
        optionally deletes the associated Python file from the filesystem.

        Args:
            id (str): Unique identifier of the stage to delete.
            remove_file (bool, optional): Whether to also delete the associated
                Python file from the filesystem. Defaults to False.

        Warning:
            - Deleting a stage that is referenced by workflow transitions may
              break the workflow flow
            - If remove_file=True, the Python file will be permanently deleted
            - This operation cannot be undone

        Copywritings:
            Delete a stage
            Deleting a stage...
        """
        with self.repositories.project.atomic() as project:
            project.delete_stage(stage_id, remove_file)

    def get_hook(self, id: str) -> HookStage | None:
        project = self.repositories.project.load()
        return project.get_hook(id)

    def get_hooks(self) -> list[HookStage]:
        project = self.repositories.project.load()
        hooks = project.get_hooks()

        sorted_hooks = sorted(hooks, key=lambda h: h.title.lower())
        return sorted_hooks

    def get_hook_by_path(self, path: str) -> HookStage | None:
        project = self.repositories.project.load()
        return project.get_hook_by_path(path)

    # Page stage methods

    def get_page_stage(self, id: str) -> PageStage | None:
        project = self.repositories.project.load()
        return project.get_page_stage(id)

    def get_page_stages(self) -> list[PageStage]:
        project = self.repositories.project.load()
        pages = project.get_page_stages()
        sorted_pages = sorted(pages, key=lambda u: u.title.lower())
        return sorted_pages

    def get_page_stage_by_path(self, path: str) -> PageStage | None:
        project = self.repositories.project.load()
        return project.get_page_stage_by_path(path)

    def run_page_stage(
        self,
        id: str,
        request: Request,
        user_jwt: str | None = None,
        page_execution_id: str | None = None,
    ):
        page = self.get_page_stage(id)
        if not page:
            raise Exception(f"Page stage with id {id} not found")

        context = PageContext(
            request=request,
            response=Response(headers={}, status=200, body=""),
            page_path=page.path,
            page_execution_id=page_execution_id,
        )

        connection = self.repositories.producer.enqueue(
            page.id, context, user_jwt=user_jwt
        )

        # First drain gets execution:started message
        start_msg = drain_until_response(
            connection, timeout=DRAIN_START_TIMEOUT_SECONDS
        )
        if not start_msg:
            connection.close()
            flask.abort(500)
            return  # unreachable, but satisfies type checker

        execution_id = None
        if isinstance(start_msg, dict) and start_msg.get("type") == "execution:started":
            execution_id = start_msg.get("executionId")

        try:
            response = normalize_response(drain_until_response(connection))

            if not response:
                flask.abort(500)
        finally:
            connection.close()

        return {
            "body": response.body,
            "status": response.status,
            "headers": response.headers,
            "executionId": execution_id,
        }

    def get_jobs(self, include_disabled_jobs: bool = False) -> list[JobStage]:
        project = self.repositories.project.load(
            include_disabled_stages=include_disabled_jobs
        )
        jobs = project.get_jobs()

        sorted_jobs = sorted(jobs, key=lambda j: j.title.lower())
        return sorted_jobs

    def get_job(self, id: str) -> JobStage | None:
        project = self.repositories.project.load()
        stage = project.get_stage(id)

        if isinstance(stage, JobStage):
            return stage

        return None

    def get_job_status(self, id: str) -> Literal["enabled", "disabled", "not_found"]:
        project = self.repositories.project.load(include_disabled_stages=True)
        stage = project.get_stage(id)

        if not isinstance(stage, JobStage):
            return "not_found"

        project = self.repositories.project.load()
        stage = project.get_stage(id)

        if isinstance(stage, JobStage):
            return "enabled"

        return "disabled"

    def create_stage(
        self,
        type: Literal["form", "page", "hook", "job", "tasklet"],
        title: str,
        file: str,
        workflow_position: tuple[int, int] = (0, 0),
        id: str | None = None,
    ) -> StageWithFile:
        """
        Create a new stage in the project workflow.

        Args:
            type: Kind of stage to create:
                - 'form': interactive form stage (collects user input)
                - 'page': interactive page stage (custom HTML/CSS/JS dashboards and tools)
                - 'hook': webhook stage (HTTP endpoint triggered externally)
                - 'job': scheduled job stage (runs periodically on a schedule)
                - 'tasklet': background script stage (processes tasks without UI)
            title (str): Display name for the new stage.
            file (str): Relative path where the stage's Python code will be stored.
                Must end with .py extension.
            workflow_position (list[int], optional): [x, y] coordinates for the
                stage's position in the visual workflow editor. Defaults to [0, 0].
            id (Optional[str], optional): Custom identifier for the stage. If None,
                a unique ID will be automatically generated.

        Copywritings:
            Create a new stage
            Creating a new stage...
        """
        if type == "form":
            stage: StageWithFile = FormStage.create(
                title, file, workflow_position=workflow_position, id=id
            )
            template = new_form_code
        elif type == "page":
            stage = PageStage.create(
                title, file, workflow_position=workflow_position, id=id
            )
            template = new_page_code
        elif type == "hook":
            stage = HookStage.create(
                title, file, workflow_position=workflow_position, id=id
            )
            template = new_hook_code
        elif type == "job":
            stage = JobStage.create(
                title, file, workflow_position=workflow_position, id=id
            )
            template = new_job_code
        elif type == "tasklet":
            stage = ScriptStage.create(
                title, file, workflow_position=workflow_position, id=id
            )
            template = new_script_code
        else:
            raise ValueError(f"Unknown stage type: {type!r}")

        self.init_code_file(stage.file, template)
        with self.repositories.project.atomic() as project:
            project.add_stage(stage)
        return stage

    def update_stage(self, id: str, changes: dict[str, Any]) -> Stage:
        """
        Update properties of an existing workflow stage.

        This method allows modification of stage metadata properties. For code
        modifications, use the specialized code editing methods instead.

        WARNING: **For code updates, use dedicated methods**:
        - `replace_code_context()` for targeted code changes (RECOMMENDED)
        - `replace_file_content()` for complete file rewrites

        Args:
            id (str): Unique identifier of the stage to update.
            changes (Dict[str, Any]): Dictionary containing the properties to update.

            **Common properties (all stage types):**
                - title (str): Display name
                - workflow_position (list): [x, y] coordinates in workflow editor

            **FormStage properties:**
                - path (str): URL path for the form
                - end_message (str): Message shown when form completes
                - start_message (str): Welcome message
                - error_message (str): Error message
                - timeout_message (str): Timeout message
                - start_button_text (str): Button text
                - auto_start (bool): Auto-start flag
                - access_control (dict): { is_public: bool, required_roles: list }
                - notification_trigger (dict): { variable_name: str, enabled: bool }

            **HookStage properties:**
                - path (str): URL path for the hook
                - enabled (bool): Enable/disable flag

            **JobStage properties:**
                - schedule (str): Cron expression (e.g., "0 9 * * *")

        Example:
            ```python
            # Update form title and access control
            controller.update_stage("form-123", {
                "title": "Customer Registration",
                "access_control": {"is_public": False, "required_roles": ["admin"]}
            })

            # Update job schedule
            controller.update_stage("job-456", {
                "schedule": "0 9 * * 1-5"  # Weekdays at 9 AM
            })
            ```

        Note:
            - Only updates stage metadata, not code content
            - Stage metadata updates are validated before being applied
            - Use specialized code editing methods for reliable code modifications

        Copywritings:
            Update properties of an existing stage
            Updating properties of an existing stage...
        """
        # Agents use prompt_content instead of code_content
        if "prompt_content" in changes:
            changes["code_content"] = changes.pop("prompt_content")

        project = self.repositories.project.load()
        stage = project.get_stage(id)

        if not stage:
            raise Exception(f"Stage with id {id} not found")

        if isinstance(stage, StageWithFile) and (
            code_content := changes.pop("code_content", None)
        ):
            temp_file = Path(mkdtemp()) / stage.file_path
            with temp_file.open("w", encoding="utf-8") as f:
                f.write(code_content)
            move(str(temp_file), Settings.root_path.joinpath(stage.file_path))

        if test_data := changes.pop("test_data", None):
            self.write_test_data(test_data)

        stage_module = project.get_stage_module(id)
        if stage_module is not None:
            module_project = stage_module.get_project()
            module_stage = module_project.get_stage(id)
            if not module_stage:
                raise Exception(f"Stage with id {id} not found in module")
            updated_stage = module_project.update_stage(module_stage, changes)
            self.repositories.project.save(module_project)
            return updated_stage

        with self.repositories.project.atomic() as project:
            atomic_stage = project.get_stage(id)
            if not atomic_stage:
                raise Exception(f"Stage with id {id} not found")
            project.update_stage(atomic_stage, changes)
        result = project.get_stage(id)
        if not result:
            raise Exception(f"Stage with id {id} not found after update")
        return result

    def list_all_stages(self) -> list[Stage]:
        """
        Retrieve all workflow stages in the current project.

        This method returns a complete list of all stages (forms, hooks, jobs,
        and tasklets) that are part of the project workflow.


        Note:
            - Only enabled stages are returned (disabled jobs are excluded)
            - Stages are returned in the order they appear in the project configuration
            - Each stage contains metadata like id, title, type, and file path

        Copywritings:
            List all workflow stages
            Retrieving all workflow stages in the project...
        """
        project = self.repositories.project.load()
        return project.workflow_stages

    # Modules
    def get_modules(self) -> list[str]:
        project = self.repositories.project.load()
        return [module.name for module in project.get_installed_modules()]

    # Login
    def get_credentials(self):
        return get_credentials()

    def get_login(self):
        headers = resolve_headers()
        if not headers:
            return {"logged": False, "reason": "NO_API_TOKEN"}
        return get_api_key_info(headers)

    def get_email(self):
        login = self.get_login()
        if login.get("logged"):
            return login.get("info", {}).get("email")
        return None

    def create_login(self, token):
        set_credentials(token)
        return self.get_login()

    def delete_login(self):
        delete_credentials()
        return self.get_login()

    # Project
    def get_project_info(self):
        headers = resolve_headers()
        if headers is None:
            flask.abort(401)
        try:
            return get_project_info(headers)
        except Exception:
            return {}

    def get_feature_flags(self):
        headers = resolve_headers()
        if headers is None:
            flask.abort(401)
        try:
            return get_feature_flags(headers)
        except Exception:
            return {}

    # access_control
    def list_access_controls(self):
        """
        List access control settings for all secured stages.

        Returns access control configuration for the home page and all forms.
        Each item includes the stage ID, title, type, and access settings.

        Copywritings:
            List access control settings
            Listing access control settings...
        """
        project = self.repositories.project.load()
        return [s.to_access_dto() for s in project.secured_stages]

    def update_access_control(
        self, id: str, is_public: bool, required_roles: list[str]
    ):
        """
        Update access control settings for a specific stage or home page.

        Modifies who can access a particular stage or the home page.
        Use 'home' as the ID to update the home page access control.

        Args:
            id (str): Stage identifier. Use 'home' for the home page, or the stage UUID for forms.
            is_public (bool): If True, anyone can access. If False, authentication is required.
            required_roles (List[str]): List of role names. User must have at least one of these roles.
                Empty list means any authenticated user can access (when is_public is False).

        Example:
            ```python
            # Make home page private, requiring 'admin' or 'manager' role
            controller.update_access_control(
                id="home",
                is_public=False,
                required_roles=["admin", "manager"]
            )

            # Make a form public
            controller.update_access_control(
                id="form-uuid-123",
                is_public=True,
                required_roles=[]
            )
            ```

        Copywritings:
            Update access control settings
            Updating access control settings...
        """
        ac_changes = [
            {"id": id, "is_public": is_public, "required_roles": required_roles}
        ]
        with self.repositories.project.atomic() as project:
            response = project.update_access_controls(ac_changes)
        return response[0] if response else None

    def update_access_controls(self, changes: list[dict[str, Any]]):
        with self.repositories.project.atomic() as project:
            response = project.update_access_controls(changes)
        return response

    def get_access_control_by_stage_id(self, id):
        project = self.repositories.project.load()
        return project.get_access_control_by_stage_id(id)

    # logs
    def get_executions(self, filter: ExecutionFilter):
        return self.execution_repository.list(filter)

    def list_executions(
        self,
        status: Optional[str] = None,
        stage_id: Optional[str] = None,
        limit: int = 20,
        offset: int = 0,
    ) -> dict:
        """
        List recent executions in the editor.

        Returns a slim summary of executions, optionally filtered by status
        and/or stage. Useful for finding the id of a running execution to
        inspect logs, fetch tasks, or stop it.

        Args:
            status (Optional[str]): Filter by execution status. One of
                "running", "failed", "finished", "abandoned". If None,
                executions of any status are returned.
            stage_id (Optional[str]): Filter by the stage that produced the
                execution. If None, executions from any stage are returned.
            limit (int): Maximum number of executions to return. Defaults to 20.
            offset (int): Number of executions to skip (for pagination).
                Defaults to 0.

        Example:
            ```python
            controller = MainController(repositories)

            # All currently running executions
            controller.list_executions(status="running")

            # Last 50 executions for a given stage
            controller.list_executions(stage_id="stage-123", limit=50)
            ```

        Copywritings:
            List recent executions
            Listing recent executions...
        """
        filter = ExecutionFilter(
            status=status,
            stage_id=stage_id,
            limit=limit,
            offset=offset,
        )
        response = self.execution_repository.list(filter)
        return {
            "executions": [
                {
                    "id": execution.id,
                    "stage_id": execution.stage_id,
                    "status": execution.status,
                    "created_at": to_utc_iso_string(execution.created_at),
                    "updated_at": (
                        to_utc_iso_string(execution.updated_at)
                        if execution.updated_at is not None
                        else None
                    ),
                }
                for execution in response.executions
            ],
            "total_count": response.total_count,
        }

    def stop_execution(self, execution_id: str):
        """
        Stop a currently running execution by its ID.

        This method terminates an in-progress stage execution. Locally, it
        sends SIGTERM to the worker process (escalating to SIGKILL if the
        process does not exit within ~2 seconds). In the web editor, it
        publishes a control message to the worker to stop the execution.

        Args:
            execution_id (str): Unique identifier of the execution to stop.

        Copywritings:
            Stop a running execution
            Stopping a running execution...
        """
        self.execution_repository.stop_execution(execution_id)

    def stop_all_executions(self):
        """
        Stop every running execution and discard queued ones.

        Order matters: queued messages are dropped first so workers cannot
        pick up a new execution between the moment we stop the current one
        and the moment we finish clearing the backlog. Then each running
        execution receives the same stop signal that the single-execution
        endpoint sends.
        """
        try:
            self.producer_repository.clear_queue()
        except Exception:
            pass

        self.execution_repository.stop_all_running()

    def get_execution_logs(self, id: str):
        """
        Retrieve execution logs for a specific execution by its ID.

        Returns the log output text from a stage execution, formatted as
        lines with event type prefixes (STDOUT/STDERR). Useful for debugging
        failed executions and understanding what happened during a run.

        Args:
            id (str): Unique identifier of the execution to retrieve logs for.


        Copywritings:
            Get execution logs for a specific execution
            Retrieving execution logs from an execution...
        """
        entries = self.execution_logs_repository.get(id)
        lines = []
        for entry in entries:
            text = entry.payload.get("text", "").rstrip("\n")
            if not text:
                continue
            prefix = entry.event.upper() if entry.event else "LOG"
            ts = (
                entry.created_at.strftime("%H:%M:%S.%f")[:-3]
                if entry.created_at
                else ""
            )
            lines.append(f"{ts} [{prefix}] {text}")

        return {
            "logs": "\n".join(lines),
            "total_entries": len(entries),
        }

    def get_execution_tasks(self, execution_id: str) -> ExecutionTasksResponse:
        """
        Retrieve task information associated with a specific execution.

        This method returns comprehensive task data for an execution, including
        the trigger task that initiated the execution and any tasks that were
        sent/created during the execution.

        Args:
            execution_id (str): Unique identifier of the execution to query.


        Note:
            - trigger_task will be None if the execution was not triggered by a task
              (e.g., manual execution, webhook, or scheduled job)
            - sent_tasks includes all tasks created during the execution lifecycle
            - This is useful for tracking workflow progression and debugging

        Copywritings:
            Get task information associated with an execution
            Retrieving task information associated with an execution...
        """
        execution = self.execution_repository.get(execution_id)

        trigger_task = None
        if isinstance(execution.context, ScriptContext):
            trigger_task = self.tasks_repository.get_by_id(execution.context.task_id)

        sent_tasks = self.tasks_repository.get_execution_sent_tasks(execution_id)

        return ExecutionTasksResponse(
            trigger_task=trigger_task,
            sent_tasks=sent_tasks,
        )

    def get_public_url(self):
        if Settings.has_public_url():
            return {"public_url": Settings.public_url}

        return {"public_url": None}

    def fail_execution(self, execution_id: str, reason: str):
        try:
            err_log = LogEntry(
                execution_id=execution_id,
                stage_id=self.execution_repository.get(execution_id).stage_id,
                created_at=datetime.datetime.now(),
                payload={"text": "[ABSTRA] Execution aborted. " + reason},
                sequence=999999,
                event="stderr",
            )
            self.execution_logs_repository.save(err_log)
        except Exception as log_error:
            AbstraLogger.error(
                f"[MainController] Failed to write abort log for execution {execution_id}: {log_error}"
            )
            AbstraLogger.capture_exception(log_error)

        self.execution_repository.set_failure_by_id(execution_id=execution_id)
        self.tasks_repository.set_locked_tasks_to_pending(execution_id)

    def _broadcast_execution_update(self, execution_id: str) -> None:
        try:
            from abstra_internals.controllers.execution.execution_stdio import (
                BroadcastController,
            )
            from abstra_internals.utils import serialize

            BroadcastController.broadcast(
                msg=serialize(
                    {
                        "type": "execution:update",
                        "payload": {"execution_id": execution_id},
                    }
                )
            )
        except Exception:
            pass

    def run_job(self, id: str, user_jwt: Optional[str] = None):
        """
        Run a job stage immediately by its ID.

        This method triggers the execution of a job stage, allowing it to run
        immediately without waiting for its scheduled time. It is useful for
        testing or manually triggering jobs.

        Args:
            id (str): Unique identifier of the job stage to run.
            user_jwt (Optional[str]): JWT token for web-editor user identification.

        Copywritings:
            Run a job for debugging
            Running a job for debugging...
        """
        status = self.get_job_status(id)
        if status == "not_found":
            raise Exception(f"Job with id {id} not found")

        if status == "disabled":
            return {"status": "disabled"}

        conn = self.repositories.producer.enqueue(
            id, context=JobContext(), user_jwt=user_jwt
        )

        hand_off = False
        try:
            start_msg = conn.recv()

            if isinstance(start_msg, str):
                start_msg = json.loads(start_msg)

            start_msg = ExecutionStartedMessage(execution_id=start_msg["executionId"])

            if WORKER_LOG_TO_QUEUE and not web_editor_uses_db():
                self._broadcast_execution_update(start_msg.execution_id)
                self.repositories.producer.consume_and_forward(conn, id)
                hand_off = True

            return {
                "ok": True,
                "execution_id": start_msg.execution_id,
            }
        finally:
            if not hand_off:
                conn.close()

    def run_hook(self, id: str, request: Request, user_jwt: Optional[str] = None):
        """
        Run a hook stage immediately by its ID for debugging.

        This simulates an HTTP request hitting the hook so you can exercise it
        without making a real network call. Construct the request you want to
        test against — only `method` is required; `query_params`, `headers`,
        and `body` default to empty when omitted.

        Args:
            id (str): Unique identifier of the hook stage to run.
            request (Request): A simulated HTTP request to feed the hook.
                Fields use snake_case (query_params, headers, method, body).
            user_jwt (Optional[str]): JWT token for web-editor user identification.

        Copywritings:
            Run a hook for debugging
            Running a hook for debugging...
        """

        hook = self.get_hook(id)
        if not hook:
            raise Exception(f"Hook with id {id} not found")

        context = HookContext(
            request=request,
            response=Response(headers={}, status=200, body=""),
        )

        connection = self.repositories.producer.enqueue(
            hook.id, context, user_jwt=user_jwt
        )
        start_msg = drain_until_response(
            connection, timeout=DRAIN_START_TIMEOUT_SECONDS
        )

        if isinstance(start_msg, str):
            try:
                start_msg = json.loads(start_msg)
            except (json.JSONDecodeError, TypeError):
                connection.close()
                flask.abort(500)
                return  # unreachable, but satisfies type checker

        if not isinstance(start_msg, dict) or "executionId" not in start_msg:
            connection.close()
            flask.abort(500)
            return  # unreachable, but satisfies type checker

        start_msg = ExecutionStartedMessage(execution_id=start_msg["executionId"])

        try:
            response = normalize_response(drain_until_response(connection))

            if not response:
                flask.abort(500)
        finally:
            connection.close()

        return {
            "status": response.status,
            "body": response.body,
            "headers": response.headers,
            "execution_id": start_msg.execution_id,
        }

    def run_tasklet(self, id: str, task_id: str, user_jwt: Optional[str] = None):
        """
        Run a tasklet stage immediately by its ID.

        This method triggers the execution of a tasklet stage, allowing it to run
        immediately without waiting for its scheduled time. It is useful for
        testing or manually triggering tasklets.

        Args:
            id (str): Unique identifier of the tasklet stage to run.
            task_id (str): ID of an existing task. You can create a task using the create_task method.
            user_jwt (Optional[str]): JWT token for web-editor user identification.

        Copywritings:
            Run a tasklet for debugging
            Running a tasklet for debugging...
        """

        script = self.get_script(id)
        if not script:
            raise Exception(f"Tasklet with id {id} not found")

        try:
            self.tasks_repository.get_by_id(task_id)
        except Exception:
            raise Exception(
                f"Task with id {task_id} not found. Please create the task first."
            )

        conn = self.repositories.producer.enqueue(
            id, context=ScriptContext(task_id=task_id), user_jwt=user_jwt
        )

        hand_off = False
        try:
            start_msg = conn.recv()

            if isinstance(start_msg, str):
                start_msg = json.loads(start_msg)

            start_msg = ExecutionStartedMessage(execution_id=start_msg["executionId"])

            if WORKER_LOG_TO_QUEUE and not web_editor_uses_db():
                self._broadcast_execution_update(start_msg.execution_id)
                self.repositories.producer.consume_and_forward(conn, id)
                hand_off = True

            return {"ok": True, "execution_id": start_msg.execution_id}
        finally:
            if not hand_off:
                conn.close()

    def _get_browser_base_url(self) -> str:
        """Get the base URL that the browser (local or remote Selenium) can use
        to reach this server.

        - Local editor: http://localhost:{port}
        - Web editor (remote Selenium): http://web-editor-{projectId}.tenants
          (reachable from the Selenium pod via K8s DNS)
        """
        project_id = os.environ.get("ABSTRA_PROJECT_ID")
        selenium_url = os.environ.get("SELENIUM_REMOTE_URL")
        if selenium_url and project_id:
            return f"http://web-editor-{project_id}.tenants"
        return f"http://localhost:{Settings.server_port}"

    def _browser_call(self, method_name: str, *args, **kwargs):
        """Dispatch a BrowserTools method call to a dedicated thread.

        Playwright has thread affinity — all operations must happen on the same
        thread that created the browser. Flask serves each request on a different
        thread, so we keep a single long-lived browser thread and proxy calls
        through a queue.
        """
        import queue as queue_mod
        import threading

        needs_init = (
            not hasattr(self, "_browser_thread")
            or self._browser_thread is None
            or not self._browser_thread.is_alive()
        )
        if needs_init:
            ready = threading.Event()
            call_queue: queue_mod.Queue = queue_mod.Queue()

            # Read token: try Flask request cookie first, fall back to file
            from abstra_internals.cloud_api import get_editor_auth_token_from_file

            try:
                editor_token = flask.request.cookies.get("editor_auth", "")
            except RuntimeError:
                editor_token = ""
            if not editor_token:
                editor_token = get_editor_auth_token_from_file()

            def _loop(token: str):
                from urllib.parse import urlparse

                from abstra_internals.agents.tools.browser import BrowserTools

                bt = BrowserTools(listen_network=True, listen_console=True)
                if token:
                    base_url = self._get_browser_base_url()
                    domain = urlparse(base_url).hostname or "localhost"
                    bt._browser_context.add_cookies(
                        [
                            {
                                "name": "editor_auth",
                                "value": token,
                                "domain": domain,
                                "path": "/",
                            }
                        ]
                    )

                def _get_iframe_frame(page_id):
                    """Get the iframe Frame for a page, or None."""
                    page = bt._get_page(page_id)
                    for frame in page.frames:
                        if frame != page.main_frame:
                            return frame
                    return None

                # Methods that should operate on the iframe content
                _iframe_methods = {
                    "get_html",
                    "get_element_html",
                    "get_text",
                    "get_page_summary",
                    "get_element_by_summary_index",
                    "click",
                    "click_element",
                    "fill",
                    "fill_element",
                    "execute_javascript",
                    "get_attribute",
                    "get_attributes",
                    "get_all_links",
                }

                ready.set()

                def _handle_iframe(name, a, kw):
                    """Handle methods that need iframe content. Returns
                    (True, result) if handled, (False, None) otherwise."""
                    if name not in _iframe_methods or not a:
                        return False, None
                    page_id = a[0]
                    frame = _get_iframe_frame(page_id)
                    if not frame:
                        return False, None

                    from abstra_internals.agents.tools.browser import (
                        _slim_element,
                        _wrap_for_safe_eval,
                    )

                    if name == "get_html":
                        return True, frame.content()

                    if name == "get_element_html":
                        selector = a[1] if len(a) > 1 else kw.get("selector", "")
                        el = frame.query_selector(selector)
                        if not el:
                            raise ValueError(f"Selector '{selector}' not found")
                        return True, el.evaluate("el => el.outerHTML")

                    if name == "get_page_summary":
                        iframe_elements = bt.extractor.extract_elements(
                            frame  # type: ignore[arg-type]
                        )
                        bt._extracted_elements[page_id] = iframe_elements
                        max_el = a[1] if len(a) > 1 else 50
                        slim = [_slim_element(e) for e in iframe_elements]
                        if len(slim) > max_el:
                            total = len(slim)
                            slim = slim[:max_el]
                            slim.append(
                                {
                                    "note": f"{total - max_el} more elements not shown (total: {total})."
                                }
                            )
                        return True, slim

                    if name == "execute_javascript":
                        script = a[1] if len(a) > 1 else kw.get("script", "")
                        return True, frame.evaluate(_wrap_for_safe_eval(script))

                    if name == "get_text":
                        selector = a[1] if len(a) > 1 else kw.get("selector", "")
                        el = frame.query_selector(selector)
                        return True, el.inner_text() if el else ""

                    if name in ("click", "click_element"):
                        # Resolve element selector from cached iframe elements,
                        # then click on the iframe frame
                        if name == "click_element":
                            index = a[1] if len(a) > 1 else kw.get("index", 0)
                            elem = bt._resolve_element(page_id, index)
                            selector = elem["selector"]
                        else:
                            selector = a[1] if len(a) > 1 else kw.get("selector")
                        if selector:
                            frame.click(selector, timeout=5000)
                        return True, None

                    if name in ("fill", "fill_element"):
                        if name == "fill_element":
                            index = a[1] if len(a) > 1 else kw.get("index", 0)
                            value = a[2] if len(a) > 2 else kw.get("value", "")
                            elem = bt._resolve_element(page_id, index)
                            selector = elem["selector"]
                        else:
                            selector = a[1] if len(a) > 1 else kw.get("selector")
                            value = a[2] if len(a) > 2 else kw.get("value", "")
                        if selector:
                            frame.fill(selector, value, timeout=5000)
                        return True, None

                    # Other iframe methods: fall through to BrowserTools
                    return False, None

                while True:
                    item = call_queue.get()
                    if item is None:
                        break
                    name, a, kw, result_q = item
                    try:
                        handled, result = _handle_iframe(name, a, kw)
                        if not handled:
                            result = getattr(bt, name)(*a, **kw)
                        result_q.put(("ok", result))
                    except Exception as e:
                        result_q.put(("error", e))

                bt.close()

            self._browser_call_queue = call_queue
            t = threading.Thread(target=_loop, args=(editor_token,), daemon=True)
            t.start()
            ready.wait()
            self._browser_thread = t

        result_q: queue_mod.Queue = queue_mod.Queue()
        self._browser_call_queue.put((method_name, args, kwargs, result_q))
        status, value = result_q.get(timeout=120)
        if status == "error":
            raise value
        return value

    def browser_open_page(self, id: str):
        """Open a project Page stage in a new browser tab by its stage ID.

        Each call opens a NEW tab and returns a unique tab_id. You can have
        multiple tabs open at once — use tab_id to target the right one.

        Returns tab_id, final url, and title. Use the tab_id in ALL subsequent
        browser_* tool calls.

        Typical workflow:
        1. browser_open_page(id) → get tab_id (new tab)
        2. browser_get_page_summary(tab_id) → see interactive elements
        3. browser_click/browser_fill → interact with elements by index
        4. browser_get_page_summary again → see updated state
        5. browser_get_console_logs → check for errors
        6. browser_close(tab_id) → close the tab when done

        Use browser_list_tabs to see all open tabs.

        Copywritings:
            Open page in browser
            Opening page in browser...
        """
        page = self.get_page_stage(id)
        if not page:
            raise Exception(f"Page with id {id} not found")

        base = self._get_browser_base_url()
        url = f"{base}/_editor/api/pages/{id}/run"
        return self._browser_call("navigate_to_url", url)

    def browser_navigate(self, url: str):
        """Navigate to any URL in a new browser tab. Each call opens a NEW tab. Returns tab_id, final url, and title. Use the tab_id in subsequent browser tool calls. Prefer browser_open_page to open project pages by ID. Close tabs with browser_close when done.

        Copywritings:
            Navigate browser to URL
            Navigating browser...
        """
        return self._browser_call("navigate_to_url", url)

    def browser_get_page_summary(self, tab_id: str):
        """List interactive elements visible on the tab (buttons, links, inputs). Each element has an index — use it with browser_click or browser_fill. Call this after any action that changes the page.

        Copywritings:
            Get page summary
            Getting page summary...
        """
        return self._browser_call("get_page_summary", tab_id)

    def browser_click(self, tab_id: str, index: int):
        """Click an interactive element by its index from browser_get_page_summary.

        Copywritings:
            Click element
            Clicking element...
        """
        return self._browser_call("click_element", tab_id, index)

    def browser_fill(self, tab_id: str, index: int, value: str):
        """Fill a form field by its index from browser_get_page_summary.

        Copywritings:
            Fill form field
            Filling form field...
        """
        return self._browser_call("fill_element", tab_id, index, value)

    def browser_get_text(self, tab_id: str, selector: str):
        """Get the inner text content of an element by CSS selector.

        Copywritings:
            Get element text
            Getting element text...
        """
        return self._browser_call("get_text", tab_id, selector)

    def browser_get_html(self, tab_id: str, selector: str):
        """Get the outer HTML of a specific element by CSS selector.

        Use page_summary to find selectors, or target well-known selectors
        like "table", "form", "main", "#content", etc.

        Args:
            tab_id (str): The tab ID from browser_open_page or run_page.
            selector (str): CSS selector for the element (e.g. "div.my-class", "#main", "table").

        Copywritings:
            Get element HTML
            Getting element HTML...
        """
        return self._browser_call("get_element_html", tab_id, selector)

    def browser_execute_javascript(self, tab_id: str, script: str):
        """Execute JavaScript on the tab and return the result. After this, call browser_get_page_summary again as the DOM may have changed.

        Copywritings:
            Execute JavaScript
            Executing JavaScript...
        """
        return self._browser_call("execute_javascript", tab_id, script)

    def browser_wait(self, tab_id: str, milliseconds: int = 1000):
        """Wait for a specified number of milliseconds. Useful after clicks or form submissions before reading page state.

        Copywritings:
            Wait
            Waiting...
        """
        return self._browser_call("wait", tab_id, milliseconds)

    def browser_get_console_logs(self, tab_id: str):
        """Get captured browser console log messages for a tab.

        Copywritings:
            Get console logs
            Getting console logs...
        """
        return self._browser_call("get_console_logs", tab_id)

    def browser_get_network_requests(self, tab_id: str):
        """Get captured network requests for a tab.

        Copywritings:
            Get network requests
            Getting network requests...
        """
        return self._browser_call("get_network_requests", tab_id)

    def browser_close(self, tab_id: str):
        """Close a browser tab. Always close tabs you no longer need to free resources.

        Copywritings:
            Close browser tab
            Closing browser tab...
        """
        return self._browser_call("close_page", tab_id)

    def browser_list_tabs(self):
        """List all open browser tabs with their tab_id, URL, and title. Use this to see which tabs are open before opening new ones.

        Copywritings:
            List browser tabs
            Listing browser tabs...
        """
        return self._browser_call("list_pages")

    def run_page(self, id: str, query_params: Optional[dict] = None):
        """Run a page, collect a full snapshot, and close it immediately.

        Opens the page in a browser, waits for it to load, collects the visible
        text, interactive elements, console logs, and network errors, then closes
        the tab automatically. The page does NOT keep running after this call.

        Use this for quick diagnostics — one call gives you everything you need
        to understand what the page rendered and whether something went wrong.

        For interactive debugging (clicking buttons, filling inputs), use
        browser_open_page instead — it keeps the tab open for interaction but
        requires you to call browser_close(tab_id) when done.

        Args:
            id (str): The page stage ID.
            query_params (dict, optional): Query parameters to append to the URL.

        Copywritings:
            Run page for debugging
            Running page for debugging...
        """
        page = self.get_page_stage(id)
        if not page:
            raise Exception(f"Page with id {id} not found")

        from urllib.parse import urlencode

        base = self._get_browser_base_url()
        url = f"{base}/_editor/api/pages/{id}/run"

        if query_params:
            url += "?" + urlencode(query_params)

        nav_result = self._browser_call("navigate_to_url", url)
        tab_id = nav_result["tab_id"]

        try:
            # Wait a moment for async rendering to settle
            self._browser_call("wait", tab_id, 1000)
            return self._collect_page_state(tab_id, url)
        finally:
            self._browser_call("close_page", tab_id)

    def _collect_page_state(self, tab_id: str, url: Optional[str] = None):
        """Internal: collect full page state for run_page."""
        page_summary = self._browser_call("get_page_summary", tab_id)
        console_logs = self._browser_call("get_console_logs", tab_id)
        text_content = self._browser_call("get_text", tab_id, "body")
        network_requests = self._browser_call("get_network_requests", tab_id)

        network_errors = [
            {
                "url": req["request"]["url"],
                "method": req["request"]["method"],
                "status": req["response"]["status"],
            }
            for req in network_requests
            if req.get("response")
            and isinstance(req["response"].get("status"), int)
            and req["response"]["status"] >= 400
        ]

        result = {
            "text_content": text_content,
            "page_summary": page_summary,
            "console_logs": console_logs,
            "network_errors": network_errors,
        }
        if url is not None:
            result["url"] = url

        return result

    def execute_code_snippet(self, code: str, title: str = "Debug Snippet"):
        """Run a Python code snippet immediately in the project's runtime environment.

        Use for: testing code before writing to files, debugging, running one-off queries,
        verifying API connections, or any auxiliary task.

        The snippet runs with full access to installed packages and project files.
        Use print() statements for output — the return value is the captured stdout/stderr.

        After testing, use the MCP file tools to write the working code to a stage file.

        Copywritings:
            Run code snippet
            Running code snippet...
        """
        from abstra_internals.repositories.models import RunSnippetMessage

        message = RunSnippetMessage.create(code=code, title=title)
        conn = self.repositories.producer.enqueue_control(message)

        try:
            result = conn.recv()
            if isinstance(result, str):
                result = json.loads(result)
            return result
        finally:
            conn.close()

    def add_and_install_requirement(self, name: str, version: str | None = None):
        """
        Add a requirement to requirements.txt and install it automatically.

        This method adds a Python package to the requirements.txt file and
        immediately installs it using pip, combining both operations in one call.

        Args:
            name (str): Name of the Python package to add and install.
            version (str, optional): Specific version to install. If not provided,
                                   installs the latest version.

        Copywritings:
            Add and install Python package
            Adding and installing Python package...
        """

        # add to requirements.txt
        requirements = RequirementsRepository.load()
        requirements.add(name, version)
        RequirementsRepository.save(requirements)

        # install the package
        try:
            installation_output = requirements.install()
        except Exception as e:
            return {
                "status": "error",
                "message": f"Installation failed: {e!s}",
                "output": [],
                "requirements": requirements.to_dict(),
            }

        if "__ABSTRA_STREAM_ERROR__" in installation_output:
            return {
                "status": "error",
                "message": f"Failed to install {name}",
                "output": installation_output,
                "requirements": requirements.to_dict(),
            }

        return {
            "status": "success",
            "message": f"Successfully added and installed {name}"
            + (f"=={version}" if version else ""),
            "output": installation_output,
            "requirements": requirements.to_dict(),
        }

    def list_linter_issues(
        self,
        type: str | None = None,
        name_pattern: str | None = None,
        limit: int = 20,
        offset: int = 0,
    ):
        """
        List linter issues found in the project codebase with filtering and pagination.

        Use this tool to discover code quality issues, security problems, and
        potential bugs. Results can be filtered by severity type and rule name.
        Use fix_issue_in_codebase to apply automatic fixes for issues that support it.

        Args:
            type (str, optional): Filter by issue severity type.
                Valid values: 'security', 'error', 'bug', 'warning', 'info'.
                If None, returns all types. Defaults to None.
            name_pattern (str, optional): Regex pattern to filter by rule name.
                Case-insensitive. Example: 'missing.*env' matches rules about
                missing environment variables. Defaults to None.
            limit (int): Maximum number of issues to return per page.
                Hard maximum of 20. Defaults to 20.
            offset (int): Number of issues to skip for pagination. Defaults to 0.

        Copywritings:
            List linter issues in the codebase
            Listing linter issues in the codebase...
        """
        import re

        MAX_LIMIT = 20
        if limit > MAX_LIMIT:
            return {"error": f"limit cannot exceed {MAX_LIMIT}, got {limit}"}

        checks = self.linter_repository.find_issues_in_codebase()

        compiled_pattern = None
        if name_pattern:
            try:
                compiled_pattern = re.compile(name_pattern, re.IGNORECASE)
            except re.error:
                return {"error": f"Invalid regex pattern: {name_pattern}"}

        flat_issues = []
        for check in checks:
            if type and check.type != type:
                continue
            if compiled_pattern and not compiled_pattern.search(check.name):
                continue
            for issue in check.issues:
                flat_issues.append(
                    {
                        "rule_name": check.name,
                        "rule_label": check.label,
                        "type": check.type,
                        "issue_label": issue.make_label(),
                        "fixes": [fix.name for fix in issue.fixes],
                    }
                )

        total = len(flat_issues)
        page = flat_issues[offset : offset + limit]

        return {
            "issues": page,
            "total": total,
            "limit": limit,
            "offset": offset,
            "has_more": offset + limit < total,
        }
