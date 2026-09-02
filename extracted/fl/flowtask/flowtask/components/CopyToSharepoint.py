"""
CopyToSharepoint — write a Pandas DataFrame directly to one SharePoint file.

Serializes the DataFrame to a temporary file (CSV / Excel / Parquet / JSON)
via ``CopyToFileBase.run()`` and then uploads it to SharePoint via the
existing ``SharepointClient.upload_files`` helper.  The temporary file is
deleted unconditionally by the base class after the upload completes or
fails.

Usage example:

.. code-block:: yaml

    CopyToSharepoint:
      credentials:
        client_id: SHAREPOINT_APP_ID
        client_secret: SHAREPOINT_APP_SECRET
        tenant_id: SHAREPOINT_TENANT_ID
      tenant: symbits
      site: Navigator-Navigator-dev
      destination:
        directory: Reports/2026-05/
        filename: monthly-report.xlsx
      format: xlsx
      sheet_name: Summary

Append rows to an existing Excel file (preserves formatting / formulas /
other sheets via openpyxl):

.. code-block:: yaml

    CopyToSharepoint:
      credentials:
        client_id: SHAREPOINT_APP_ID
        client_secret: SHAREPOINT_APP_SECRET
        tenant_id: SHAREPOINT_TENANT_ID
      tenant: symbits
      site: Navigator-Navigator-dev
      destination:
        directory: Reports/2026-05/
        filename: ventas.xlsx
      format: xlsx
      sheet_name: Ventas
      mode: append                 # write (default) | append
      create_if_missing: true      # create the file if it does not exist
      on_schema_mismatch: error    # error (default) | reorder | ignore

Target a user's personal OneDrive instead of a SharePoint site library by
adding a ``drive`` block (works in both write and append modes):

.. code-block:: yaml

    CopyToSharepoint:
      credentials: { client_id: ..., client_secret: ..., tenant_id: ... }
      tenant: symbits
      drive:
        type: onedrive             # site (default) | onedrive
        user: someone@symbits.com  # UPN, user-id, or "me" (delegated auth)
      destination:
        directory: Documents/Reports/
        filename: ventas.xlsx
      format: xlsx
      mode: append
"""
from __future__ import annotations

import shutil
import tempfile
from collections.abc import Callable
from datetime import datetime
from pathlib import Path
from typing import Any, List, Optional

import pandas

from ..exceptions import ComponentError, DataNotFound, FileError, FileNotFound
from ..interfaces.Sharepoint import SharepointClient
from ..interfaces.copy_to_file_base import CopyToFileBase

_APPEND_MODES = {"write", "append"}
_SCHEMA_POLICIES = {"error", "reorder", "ignore"}


class CopyToSharepoint(SharepointClient, CopyToFileBase):
    """Upload a DataFrame to a single SharePoint file in CSV / Excel / Parquet / JSON format.

    Inherits from ``SharepointClient`` (left) and ``CopyToFileBase`` (right),
    mirroring the inheritance order of
    ``UploadToSharepoint(SharepointClient, UploadToBase)``.

    The v1 implementation uses the tempfile fallback path:
    ``CopyToFileBase.run()`` falls back to ``_put_from_tempfile()`` when
    ``_put_from_bytes()`` raises ``NotImplementedError`` (the default).

    Properties (in addition to CopyToFileBase write options):

    .. table::
        :widths: auto

        +------------------+-----+-------------------------------------------------------+
        | credentials      | Yes | SharePoint credentials (app or user auth).            |
        +------------------+-----+-------------------------------------------------------+
        | tenant           | Yes | SharePoint tenant name (e.g. ``symbits``).            |
        +------------------+-----+-------------------------------------------------------+
        | site             | Yes | SharePoint site name.                                 |
        +------------------+-----+-------------------------------------------------------+
        | destination      | Yes | Dict ``{ directory: <path>, filename: <name> }``      |
        |                  |     | — single file only in v1.                             |
        +------------------+-----+-------------------------------------------------------+
        | format           | No  | Output format: ``csv`` (default), ``xlsx``,           |
        |                  |     | ``parquet``, ``json``.                                |
        +------------------+-----+-------------------------------------------------------+

    Note:
        Single-file only in v1. For multi-file uploads use ``UploadToSharepoint``.
    """

    _version = "1.0.0"

    _credentials: dict = {
        "client_id": str,
        "client_secret": str,
        "tenant_id": str,
        "username": str,
        "password": str,
        "tenant": str,
        "site": str,
    }

    def __init__(
        self,
        loop=None,
        job: Optional[Callable] = None,
        stat: Optional[Callable] = None,
        **kwargs,
    ) -> None:
        """Initialise, forwarding kwargs through the MRO chain.

        Args:
            loop: Unused; accepted for API parity.
            job: Optional job callable.
            stat: Optional stat callable.
            **kwargs: All component configuration, including ``credentials``,
                      ``tenant``, ``site``, ``destination``, ``format``,
                      and all CopyToFileBase write options.
        """
        self.url: Optional[str] = None
        self.context = None
        self.destination_dir: str = "Shared Documents/"
        self.destination_filename: Optional[str] = None

        # Append-mode options (consumed here so they are not forwarded as
        # arbitrary FlowComponent attributes).
        self.mode: str = (kwargs.pop("mode", "write") or "write").strip().lower()
        if self.mode not in _APPEND_MODES:
            raise ComponentError(
                f"CopyToSharepoint: invalid mode {self.mode!r}. "
                f"Expected one of {sorted(_APPEND_MODES)}."
            )
        self.create_if_missing: bool = bool(kwargs.pop("create_if_missing", True))
        self.on_schema_mismatch: str = (
            kwargs.pop("on_schema_mismatch", "error") or "error"
        ).strip().lower()
        if self.on_schema_mismatch not in _SCHEMA_POLICIES:
            raise ComponentError(
                f"CopyToSharepoint: invalid on_schema_mismatch "
                f"{self.on_schema_mismatch!r}. Expected one of "
                f"{sorted(_SCHEMA_POLICIES)}."
            )
        self.write_header_if_empty: bool = bool(
            kwargs.pop("write_header_if_empty", True)
        )
        super().__init__(loop=loop, job=job, stat=stat, **kwargs)

    async def start(self, **kwargs) -> bool:
        """Prepare the component, resolving the destination path and filename.

        Args:
            **kwargs: Forwarded to super().start().

        Returns:
            ``True`` on success.

        Raises:
            ComponentError: When ``destination.filename`` is missing or a list.
        """
        await super().start(**kwargs)

        dest = getattr(self, "destination", None) or {}

        # Resolve directory prefix
        directory = dest.get("directory", "Shared Documents") or "Shared Documents"
        if "{" in directory:
            directory = self.mask_replacement(directory)
        if not directory.endswith("/"):
            directory += "/"
        self.destination_dir = directory

        # Resolve filename — single file only
        filename = dest.get("filename")
        if isinstance(filename, list):
            raise ComponentError(
                "CopyToSharepoint: single-file only in v1. "
                "Use UploadToSharepoint for multi-file uploads."
            )
        if not filename:
            raise ComponentError(
                "CopyToSharepoint: missing destination.filename. "
                "Provide `destination: { directory: <path>, filename: <name> }`."
            )
        if isinstance(filename, str) and "{" in filename:
            filename = self.mask_replacement(filename)
        self.destination_filename = filename
        return True

    async def _put_from_tempfile(self, path: Path) -> None:
        """Upload the serialised DataFrame tempfile to SharePoint.

        Called by ``CopyToFileBase.run()`` after writing the DataFrame to a
        temporary file.  The base class deletes the tempfile unconditionally
        after this method returns.

        Args:
            path: Path to the temporary file containing the serialised
                  DataFrame.

        Raises:
            FileError: When ``upload_files`` reports errors.
        """
        async with self.connection():
            await self.verify_sharepoint_access()
            if not self.context:
                self.context = self.get_context(self.url)
            result = await self.upload_files(
                filenames=[path],
                destination=self.destination_dir,
                destination_filenames=[self.destination_filename],
            )

        # Validate the result — upload_files may return a dict with an "errors" key
        if isinstance(result, dict) and result.get("errors"):
            raise FileError(
                f"CopyToSharepoint: upload errors for "
                f"'{self.destination_filename}' → '{self.destination_dir}': "
                f"{result['errors']}"
            )

        self.add_metric(
            "SHAREPOINT_UPLOADED",
            {self.destination_filename: self.destination_dir},
        )
        self._logger.info(
            "CopyToSharepoint: uploaded %s to %s",
            self.destination_filename,
            self.destination_dir,
        )

    # ------------------------------------------------------------------
    # Append mode (xlsx-only in v1, openpyxl strategy)
    # ------------------------------------------------------------------

    async def run(self) -> pandas.DataFrame:
        """Dispatch to write (default) or append behaviour.

        ``mode: write`` keeps the original ``CopyToFileBase.run()`` pipeline
        (full-file replace). ``mode: append`` downloads the existing Excel
        file, appends the input rows with openpyxl (preserving formatting,
        formulas and other sheets) and re-uploads it.

        Returns:
            The input ``pandas.DataFrame`` unchanged.
        """
        if self.mode != "append":
            return await super().run()
        return await self._run_append()

    async def _run_append(self) -> pandas.DataFrame:
        """Download → append rows (openpyxl) → re-upload the target xlsx.

        Raises:
            DataNotFound: When ``self.data`` is ``None`` or empty.
            ComponentError: When ``format`` is not ``xlsx``.
            FileNotFound: When the target is missing and
                ``create_if_missing`` is ``False``.
            FileError: On schema mismatch (``on_schema_mismatch='error'``)
                or upload failure.

        Note:
            The download-modify-upload cycle is NOT atomic. Concurrent appends
            to the same file may lose rows; serialise them at the task level.
        """
        df = getattr(self, "data", None)
        if df is None or (hasattr(df, "empty") and df.empty):
            raise DataNotFound(f"{self.__class__.__name__}: no input DataFrame")
        if self.format != "xlsx":
            raise ComponentError(
                "CopyToSharepoint: mode=append only supports format=xlsx in v1."
            )

        async with self.connection():
            await self.verify_sharepoint_access()
            if not self.context:
                self.context = self.get_context(self.url)

            tmpdir = Path(tempfile.mkdtemp(prefix="appendto-sharepoint-"))
            try:
                existing = await self._download_existing(tmpdir)

                if existing is None:
                    if not self.create_if_missing:
                        raise FileNotFound(
                            f"CopyToSharepoint: target "
                            f"'{self.destination_filename}' not found in "
                            f"'{self.destination_dir}' and create_if_missing=False."
                        )
                    # Create a fresh workbook from the DataFrame and upload it.
                    out = tmpdir / self.destination_filename
                    df.to_excel(
                        out, sheet_name=self.sheet_name, index=self.index
                    )
                    await self._upload(out)
                    self.add_metric(
                        "SHAREPOINT_APPENDED",
                        {
                            self.destination_filename: self.destination_dir,
                            "rows": int(len(df)),
                            "created": True,
                        },
                    )
                    self._logger.info(
                        "CopyToSharepoint: created %s in %s (%d rows)",
                        self.destination_filename,
                        self.destination_dir,
                        len(df),
                    )
                else:
                    appended = self._append_rows(existing, df)
                    await self._upload(existing)
                    self.add_metric(
                        "SHAREPOINT_APPENDED",
                        {
                            self.destination_filename: self.destination_dir,
                            "rows": int(appended),
                            "created": False,
                        },
                    )
                    self._logger.info(
                        "CopyToSharepoint: appended %d rows to %s in %s",
                        appended,
                        self.destination_filename,
                        self.destination_dir,
                    )
            finally:
                shutil.rmtree(tmpdir, ignore_errors=True)

        self._result = df
        return df

    async def _download_existing(self, tmpdir: Path) -> Optional[Path]:
        """Download the append target into ``tmpdir`` if it exists.

        Reuses ``SharepointClient.file_search`` + ``download_found_files``
        (the same primitives ``CopyFromSharepoint`` relies on).

        Returns:
            Path to the downloaded file, or ``None`` when the target does not
            exist in SharePoint.
        """
        # file_search reads these; directory is where downloads land.
        self._srcfiles = [
            {
                "filename": self.destination_filename,
                "directory": self.destination_dir,
                "recursive": False,
            }
        ]
        self.directory = str(tmpdir)
        self._filenames = None  # keep original names on download

        try:
            found = await self.file_search()
        except (FileError, FileNotFound):
            # file_search raises FileError when nothing matches.
            return None

        if not found:
            return None

        await self.download_found_files(found)

        candidates = [p for p in tmpdir.iterdir() if p.is_file()]
        if not candidates:
            return None
        exact = [p for p in candidates if p.name == self.destination_filename]
        return exact[0] if exact else candidates[0]

    def _append_rows(self, path: Path, df: pandas.DataFrame) -> int:
        """Append ``df`` rows to ``path`` (xlsx) with openpyxl, in place.

        Preserves existing formatting, formulas and other sheets. Reconciles
        columns against the existing header per ``on_schema_mismatch``.

        Returns:
            Number of data rows appended.
        """
        from openpyxl import load_workbook

        if self.index:
            df = df.reset_index()

        wb = load_workbook(path)  # data_only=False → keep formulas/styles
        if self.sheet_name in wb.sheetnames:
            ws = wb[self.sheet_name]
        else:
            ws = wb.create_sheet(self.sheet_name)

        columns = [str(c) for c in df.columns]

        # Is there a usable header row already? A freshly-created/empty sheet
        # still reports max_row == 1 with all-None cells, so check values.
        has_data = ws.max_row >= 1 and any(
            cell.value is not None for cell in ws[1]
        )

        # Use an explicit row cursor with ws.cell() rather than ws.append():
        # append() writes after max_row, which on an empty sheet (phantom
        # max_row == 1) would leave a blank first row.
        if has_data:
            existing_header = [
                cell.value for cell in ws[1] if cell.value is not None
            ]
            df = self._reconcile_schema(df, existing_header)
            columns = [str(c) for c in df.columns]
            next_row = ws.max_row + 1
        else:
            next_row = 1
            if self.write_header_if_empty:
                for col_idx, col in enumerate(columns, start=1):
                    ws.cell(row=next_row, column=col_idx, value=col)
                next_row += 1

        appended = 0
        for row in df.itertuples(index=False, name=None):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=next_row, column=col_idx, value=self._coerce(value))
            next_row += 1
            appended += 1

        wb.save(path)
        return appended

    def _reconcile_schema(
        self, df: pandas.DataFrame, existing_header: List[Any]
    ) -> pandas.DataFrame:
        """Align ``df`` columns to the existing sheet header.

        Policy (``on_schema_mismatch``):
          - ``error``  : raise if column *sets* differ (missing/extra).
          - ``reorder``: reorder to the existing header; missing → blank,
            extra columns dropped.
          - ``ignore`` : append positionally, leaving ``df`` untouched.
        """
        header = [str(h) for h in existing_header]
        df_cols = [str(c) for c in df.columns]

        if self.on_schema_mismatch == "ignore":
            return df

        missing = [h for h in header if h not in df_cols]   # in sheet, not in df
        extra = [c for c in df_cols if c not in header]      # in df, not in sheet

        if self.on_schema_mismatch == "error":
            if missing or extra:
                raise FileError(
                    "CopyToSharepoint: append schema mismatch for "
                    f"'{self.destination_filename}'. "
                    f"Missing in data: {missing}; unexpected in data: {extra}. "
                    "Set on_schema_mismatch='reorder' or 'ignore' to override."
                )
            # Same set — reorder to the existing header order.
            return df.reindex(columns=header)

        # reorder: align to header, fill missing with blanks, drop extras.
        return df.reindex(columns=header)

    @staticmethod
    def _coerce(value: Any) -> Any:
        """Coerce a pandas/numpy scalar into an openpyxl-friendly Python value.

        NaN/NaT → ``None``; ``pandas.Timestamp`` → ``datetime``; numpy scalars
        → native Python via ``.item()``.
        """
        if value is None:
            return None
        try:
            if pandas.isna(value):
                return None
        except (TypeError, ValueError):
            pass
        if isinstance(value, pandas.Timestamp):
            return value.to_pydatetime()
        if isinstance(value, datetime):
            return value
        item = getattr(value, "item", None)
        if callable(item):
            try:
                return item()
            except Exception:  # pragma: no cover - defensive
                return value
        return value

    async def _upload(self, path: Path) -> None:
        """Upload ``path`` to the configured destination (replace), with checks."""
        result = await self.upload_files(
            filenames=[path],
            destination=self.destination_dir,
            destination_filenames=[self.destination_filename],
        )
        if isinstance(result, dict) and result.get("errors"):
            raise FileError(
                f"CopyToSharepoint: upload errors for "
                f"'{self.destination_filename}' → '{self.destination_dir}': "
                f"{result['errors']}"
            )

    async def close(self) -> None:
        """Close the SharePoint connection gracefully.

        Wraps ``SharepointClient.close()`` and suppresses errors so that
        cleanup does not mask the primary exception from ``run()``.
        """
        try:
            await super().close()
        except Exception:
            self._logger.warning(
                "CopyToSharepoint: error during close()", exc_info=True
            )
