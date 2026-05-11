from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from img2table._validation import ValidationError
from img2table.document.image import Image
from img2table.ocr import TesseractOCR
from img2table.tables.extraction import BBox


def test_validators() -> None:
    with pytest.raises(ValidationError):
        Image(src=1)  # ty:ignore[invalid-argument-type]

    with pytest.raises(ValidationError):
        Image(src="img", detect_rotation=3)  # ty:ignore[invalid-argument-type]


def test_load_image() -> None:
    # Load from path
    img_from_path = Image(src="test_data/test.png")

    # Load from bytes
    with Path("test_data/test.png").open("rb") as f:
        img_from_bytes = Image(src=f.read())

    # Load from BytesIO
    with Path("test_data/test.png").open("rb") as f:
        img_from_bytesio = Image(src=BytesIO(f.read()))

    assert img_from_path.file_bytes == img_from_bytes.file_bytes == img_from_bytesio.file_bytes

    assert next(iter(img_from_path.images)).shape == (417, 1365, 3)


def test_blank_image(mock_tesseract) -> None:  # noqa: ANN001, ARG001
    ocr = TesseractOCR()
    img = Image(src="test_data/blank.png", detect_rotation=True)

    result = img.extract_tables(
        ocr=ocr, implicit_rows=True, borderless_tables=True, min_confidence=50
    )

    assert result == []


def test_blank_no_ocr() -> None:
    img = Image(src="test_data/blank.png", detect_rotation=True)

    result = img.extract_tables(implicit_rows=True, borderless_tables=True, min_confidence=50)

    assert result == []


def test_image_tables(mock_tesseract) -> None:  # noqa: ANN001, ARG001
    ocr = TesseractOCR()
    img = Image(src="test_data/test.png", detect_rotation=True)
    img_height, img_width = img.images[0].shape[:2]

    result = img.extract_tables(ocr=ocr, implicit_rows=True, min_confidence=50)

    assert len(result) == 2

    assert result[0].title is None
    assert result[0].bbox == BBox(x1=36, y1=22, x2=770, y2=328)
    assert result[0].bbox.relative.x1 == pytest.approx(36 / img_width)
    assert result[0].bbox.relative.y1 == pytest.approx(22 / img_height)
    assert result[0].bbox.relative.x2 == pytest.approx(770 / img_width)
    assert result[0].bbox.relative.y2 == pytest.approx(328 / img_height)
    assert len(result[0].content) == 6
    assert len(result[0].content[0]) == 3

    assert result[1].title is None
    assert result[1].bbox == BBox(x1=962, y1=22, x2=1155, y2=124)
    assert result[1].bbox.relative.x1 == pytest.approx(962 / img_width)
    assert result[1].bbox.relative.y1 == pytest.approx(22 / img_height)
    assert result[1].bbox.relative.x2 == pytest.approx(1155 / img_width)
    assert result[1].bbox.relative.y2 == pytest.approx(124 / img_height)
    assert len(result[1].content) == 2
    assert len(result[1].content[0]) == 2


def test_no_ocr() -> None:
    img = Image(src="test_data/dark.png", detect_rotation=True)
    img_height, img_width = img.images[0].shape[:2]

    result = img.extract_tables(implicit_rows=True, min_confidence=50)

    assert len(result) == 1

    assert result[0].title is None
    assert result[0].bbox == BBox(x1=40, y1=37, x2=834, y2=526)
    assert result[0].bbox.relative.x1 == pytest.approx(40 / img_width)
    assert result[0].bbox.relative.y1 == pytest.approx(37 / img_height)
    assert result[0].bbox.relative.x2 == pytest.approx(834 / img_width)
    assert result[0].bbox.relative.y2 == pytest.approx(526 / img_height)
    assert len(result[0].content) == 19
    assert len(result[0].content[0]) == 5


def test_image_excel(mock_tesseract) -> None:  # noqa: ANN001, ARG001
    ocr = TesseractOCR()
    img = Image(src="test_data/test.png", detect_rotation=True)

    result = img.to_xlsx(dest=BytesIO(), ocr=ocr, implicit_rows=True, min_confidence=50)

    expected = load_workbook(filename="test_data/expected.xlsx")
    result_wb = load_workbook(filename=result)

    for idx, ws in enumerate(result_wb.worksheets):
        assert ws.title == expected.worksheets[idx].title
        assert list(ws.values) == list(expected.worksheets[idx].values)
